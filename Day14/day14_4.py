

# 쇼핑 검색 후에 엑셀파일로 저장
import json
import urllib.parse
import urllib.request
from openpyxl import Workbook

def 쇼핑검색( ) :
    display = input("검색결과수 : ")
    client_id = "w9HPzgkd7qUCVergunhj"  # 신청한 클라이언트의 아이디
    client_secret = "ltRrRXkmKV"  # 신청한 클라이언트의 비밀번호
    # url설정
    url = "https://openapi.naver.com/v1/search/shop.json"       # 데이터 가져오기 위한 url
    option = "&display="+display + "&sort=sim"                  # 옵션
    query = "?query=" + urllib.parse.quote( input("shop : 검색 : ")) # 조건
    url_query = url + query + option
    #요청
    request = urllib.request.Request(url_query)                 # 해당 url 요청
    request.add_header("X-Naver-Client-Id", client_id)          # 요청 클라이언 아이디
    request.add_header("X-Naver-Client-Secret", client_secret)   # 요청 클라이언 비밀번호
    #응답
    respone =urllib.request.urlopen( request )  # 요청에 대한 응답
    rescode = respone.getcode()
    if( rescode == 200 ) :      # 요청결과가 200 이면 요청 성공
        response_body = respone.read()  # 결과 읽어오기
    # 데이터 처리
    text = response_body.decode("utf-8")
    jsontext = json.loads(text)

    titlelist = [ ]
    brandlist = [ ]
    lpricelist = [ ]
    linklist = [ ]
    for i in jsontext['items'] :
        titlelist.append( i['title'] )
        brandlist.append(i['brand'])
        lpricelist.append(i['lprice'])
        linklist.append(i['link'])
    # 엑셀로 저장
    wb = Workbook( )                # 엑셀 객체 클래스
    filename = "shopresult.xlsx"    # 엑셀파일 이름 정하기
    sheet = wb.active               # 시트 활성화
    sheet.title = "shop"            # 시트 이름 정하기

    for row in range( 1 , (int(display)+1)) : # row는 1부터 검색결과수까지 반복
        sheet.cell( row = row , column=1 ).value = row
        sheet.cell(row=row, column=2).value = titlelist[ row-1 ]
        sheet.cell(row=row, column=3).value = brandlist[ row-1 ]
        sheet.cell(row=row, column=4).value = lpricelist[ row-1 ]
        sheet.cell(row=row, column=5).value = linklist[ row-1 ]

    wb.save( filename=filename )

쇼핑검색()









