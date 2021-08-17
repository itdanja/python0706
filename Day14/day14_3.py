import json
import re
import urllib.parse
import urllib.request
from openpyxl import Workbook
#검색 함수
def 네이버검색( 카테고리 , 검색결과수 ) :
    # 1. 클라이언 정보
    client_id = "w9HPzgkd7qUCVergunhj"  # 신청한 클라이언트의 아이디
    client_secret = "ltRrRXkmKV"  # 신청한 클라이언트의 비밀번호
    # 2. 도서검색 url
    url = "https://openapi.naver.com/v1/search/"+카테고리+".json"
    # 3. 옵션
    if 카테고리 == "news" :
        option = "&display=" + 검색결과수 + "&sort=date"  # display="검색갯수"
    elif 카테고리 == "shop":
        option = "&display=" + 검색결과수 + "&sort=sim"  # display="검색갯수"
    else :
        option = "&display=" + 검색결과수 + "&sort=count"  # display="검색갯수"
    # 4. 조건
    query = "?query=" + urllib.parse.quote(input(카테고리+" 검색 :") )
    # 5. url 합치기
    url_query = url + query + option
##############################################################
    # api 검색 요청 개체 설정
    request = urllib.request.Request(url_query)
    request.add_header("X-Naver-Client-Id", client_id)
    request.add_header("X-Naver-Client-Secret", client_secret)
    # 검색 요청
    response = urllib.request.urlopen(request)
    rescode = response.getcode()
    if (rescode == 200):
        response_body = response.read()
    else:
        print("Error Code:" + rescode)
##############################################################
    text = response_body.decode("utf-8")
    jsontext = json.loads(text)
    list = [ ] # 검색 결과 담을 리스트

    for i in jsontext['items'] :
        # 정규표현식을 이용한 <br> 태그 제거
        title = re.sub("<.+?>" , "" , i["title"] )
        print( title )
        list.append( title )
################엑셀 파일########################
    확인 = input(" 검색 결과를 엑셀로 저장할까요? [ Y / N ]")
    if 확인 =="Y" :
        wb = Workbook( )  # 엑셀 저장
        파일명 = "검색결과.xlsx" # 엑셀 파일명
        시트1 = wb.active # 데이터를 넣을 시트를 활성화
        시트1.title = "Sheet1"# 시트 이름 정하기
        # 셀[ cell : 한칸 ]
        for 행 in range( 1 , (int(검색결과수)+1) ) :
            시트1.cell( row=행 , column=1 ).value = 행
            시트1.cell( row=행 , column=2 ).value = list[ (행-1) ]
        # 엑셀 저장
        wb.save( filename=파일명 )
###########################################################
while True :
    print(" 검색[NaverAPI] 프로그램 ")
    ch = input(" 카테고리 [ 1.뉴스 2.쇼핑 3.도서 ] 0:종료 ")
    if ch == "1" :
        카테고리 ="news"
        검색결과수 =input("검색결과수: ")
        네이버검색( 카테고리 , 검색결과수 )
    if ch == "2" :
        카테고리 = "shop"
        검색결과수 = input("검색결과수: ")
        네이버검색(카테고리, 검색결과수)
    if ch == "3" :
        카테고리 = "book"
        검색결과수 = input("검색결과수: ")
        네이버검색(카테고리, 검색결과수)





















